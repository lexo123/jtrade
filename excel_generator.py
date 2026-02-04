"""
Excel Template Generator
Reads template.xls, applies specified cell changes, and generates new Excel files.
"""

import os
from datetime import datetime
from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from copy import copy
from pathlib import Path
import xlrd
import subprocess
import shutil


class ExcelTemplateGenerator:
    def __init__(self, template_path):
        """
        Initialize with template file path.
        
        Args:
            template_path (str): Path to the template Excel file (.xls or .xlsx)
        """
        self.template_path = template_path
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Template file not found: {template_path}")
    
    def _load_workbook(self, file_path):
        """
        Load workbook from either .xls or .xlsx format.
        Converts .xls to .xlsx preserving all formatting.
        
        Args:
            file_path (str): Path to Excel file
            
        Returns:
            openpyxl Workbook object
        """
        if file_path.lower().endswith('.xls'):
            # Use pandas to read and convert .xls to .xlsx with formatting preserved
            try:
                import pandas as pd
                import tempfile
                
                # Read all sheets from .xls file
                xls_file = pd.ExcelFile(file_path)
                
                # Create a temporary .xlsx file
                temp_xlsx = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
                temp_xlsx_path = temp_xlsx.name
                temp_xlsx.close()
                
                # Write to temporary .xlsx (pandas preserves basic formatting)
                with pd.ExcelWriter(temp_xlsx_path, engine='openpyxl') as writer:
                    for sheet_name in xls_file.sheet_names:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
                        df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
                
                # Load the temporary .xlsx file
                wb = openpyxl_load_workbook(temp_xlsx_path)
                
                # Store temp file path for cleanup later
                self._temp_file = temp_xlsx_path
                
                return wb
            except ImportError:
                # Fallback if pandas not available
                print("Warning: pandas not available, using basic conversion (some formatting may be lost)")
                return self._load_workbook_basic(file_path)
        else:
            # Load as .xlsx directly
            return openpyxl_load_workbook(file_path)
    
    def _load_workbook_basic(self, file_path):
        """
        Basic fallback conversion from .xls to .xlsx (minimal formatting)
        """
        xls_book = xlrd.open_workbook(file_path)
        xls_sheet = xls_book.sheet_by_index(0)
        
        xlsx_wb = Workbook()
        xlsx_ws = xlsx_wb.active
        
        # Copy all cells from xls to xlsx
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_idx, col_idx)
                if cell_value:
                    xlsx_ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)
        
        return xlsx_wb
    
    def _modify_excel_in_place(self, excel_path, company_name, sakadastro, address, invoice_number, changes, items):
        """
        Modify Excel file in-place using XML manipulation to preserve images.
        This modifies the worksheet XML directly without losing image references.
        """
        try:
            from zipfile import ZipFile, ZIP_DEFLATED
            import tempfile
            import xml.etree.ElementTree as ET

            # Create temp directory
            temp_dir = tempfile.mkdtemp()

            # Extract the Excel file
            with ZipFile(excel_path, 'r') as z:
                z.extractall(temp_dir)

            # Find the correct worksheet XML file
            import glob
            sheet_xml_paths = glob.glob(os.path.join(temp_dir, 'xl', 'worksheets', '*.xml'))
            if not sheet_xml_paths:
                raise Exception("No worksheet XML files found in Excel archive")

            # Use the first worksheet (typically sheet1.xml)
            sheet_xml_path = sheet_xml_paths[0]
            
            # Parse XML
            ET.register_namespace('', 'http://schemas.openxmlformats.org/spreadsheetml/2006/main')
            ET.register_namespace('r', 'http://schemas.openxmlformats.org/officeDocument/2006/relationships')
            
            tree = ET.parse(sheet_xml_path)
            root = tree.getroot()
            
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            
            # Find sheet data
            sheet_data = root.find('main:sheetData', ns)
            if sheet_data is None:
                raise Exception("Could not find sheetData in worksheet")
            
            # Create a simple cell value setter
            def set_cell(sheet_data, cell_ref, value):
                """Set a cell value in the XML"""
                # Parse cell reference (e.g., "A12")
                import re
                match = re.match(r'([A-Z]+)(\d+)', cell_ref)
                if not match:
                    return
                
                col_letters, row_num = match.groups()
                row_num = int(row_num)
                
                # Find or create row
                row = None
                for r in sheet_data.findall('main:row', ns):
                    if int(r.get('r')) == row_num:
                        row = r
                        break
                
                if row is None:
                    row = ET.SubElement(sheet_data, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}row')
                    row.set('r', str(row_num))
                
                # Find or create cell
                cell_address = f'{col_letters}{row_num}'
                cell = None
                for c in row.findall('main:c', ns):
                    if c.get('r') == cell_address:
                        cell = c
                        break
                
                if cell is None:
                    cell = ET.SubElement(row, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}c')
                    cell.set('r', cell_address)
                
                # Skip if cell has a formula - preserve template formulas
                if cell.find('main:f', ns) is not None:
                    return
                
                # Remove existing children to avoid keeping shared references
                for child in list(cell):
                    cell.remove(child)

                # Set value correctly depending on type.
                # For numbers, use a plain <v> element. For strings, use inlineStr
                if value is None:
                    return
                # Datetime: write as Excel serial number so Excel will display formatted date/time
                if isinstance(value, datetime):
                    # Excel epoch: 1899-12-30
                    epoch = datetime(1899, 12, 30)
                    delta = value - epoch
                    serial = delta.total_seconds() / 86400.0
                    # Write numeric value
                    cell.attrib.pop('t', None)
                    v = ET.SubElement(cell, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    # keep full precision
                    v.text = repr(serial)
                # Numeric types (int/float)
                elif isinstance(value, (int, float)):
                    cell.attrib.pop('t', None)
                    v = ET.SubElement(cell, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                    v.text = str(value)
                else:
                    # Use inline string to avoid messing with sharedStrings.xml
                    cell.set('t', 'inlineStr')
                    is_el = ET.SubElement(cell, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is')
                    t_el = ET.SubElement(is_el, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t')
                    t_el.text = str(value)
            
            # Set required fields
            set_cell(sheet_data, 'D4', datetime.now())
            set_cell(sheet_data, 'A12', company_name)
            set_cell(sheet_data, 'A13', sakadastro)
            set_cell(sheet_data, 'A14', address)
            set_cell(sheet_data, 'D5', invoice_number)
            
            # Helper to clear cached value from formula cells so Excel recalculates
            def clear_formula_cache(sheet_data, cell_ref):
                """Remove <v> element from formula cells to force recalculation"""
                import re
                match = re.match(r'([A-Z]+)(\d+)', cell_ref)
                if not match:
                    return
                col_letters, row_num = match.groups()
                row_num = int(row_num)
                cell_address = f'{col_letters}{row_num}'
                for r in sheet_data.findall('main:row', ns):
                    if int(r.get('r')) == row_num:
                        for c in r.findall('main:c', ns):
                            if c.get('r') == cell_address:
                                # If this cell has a formula, remove cached <v>
                                if c.find('main:f', ns) is not None:
                                    v_elem = c.find('main:v', ns)
                                    if v_elem is not None:
                                        c.remove(v_elem)
                                break
                        break
            
            # Set items
            start_row = 17
            for i, item in enumerate(items):
                if i >= 8:
                    break
                row = start_row + i
                if isinstance(item, dict):
                    set_cell(sheet_data, f'A{row}', item.get('type', ''))
                    set_cell(sheet_data, f'B{row}', item.get('quantity', ''))
                    set_cell(sheet_data, f'C{row}', item.get('price', ''))
                    # Clear cached value from D row so formula recalculates
                    clear_formula_cache(sheet_data, f'D{row}')
            
            # Also clear D36 (likely a sum formula) to force recalculation
            clear_formula_cache(sheet_data, 'D36')
            
            # Save modified XML
            tree.write(sheet_xml_path, encoding='utf-8', xml_declaration=True)
            
            # Re-zip the file
            with ZipFile(excel_path, 'w', ZIP_DEFLATED) as z:
                for root_dir, dirs, files in os.walk(temp_dir):
                    for file in files:
                        file_path = os.path.join(root_dir, file)
                        arcname = os.path.relpath(file_path, temp_dir)
                        z.write(file_path, arcname)
            
            # Clean up temp directory
            shutil.rmtree(temp_dir)
            
            print(f"Excel file modified in-place: {excel_path}")
            
        except Exception as e:
            print(f"Error modifying Excel in-place: {e}")
            import traceback
            traceback.print_exc()
    
    def _preserve_images_in_copy(self, template_path, output_path):
        """
        Manually copy images, drawings, and relationships from template to output.
        This is a workaround for openpyxl not preserving images and drawing references.
        """
        try:
            from zipfile import ZipFile, ZIP_DEFLATED
            import tempfile
            
            # Read from template
            with ZipFile(template_path, 'r') as template_zip:
                # Find all files that need to be preserved
                files_to_preserve = []
                
                # Media/image files
                for f in template_zip.namelist():
                    if f.startswith('xl/media/') or f.startswith('xl/drawings/'):
                        files_to_preserve.append(f)
                
                if not files_to_preserve:
                    return  # No images/drawings to preserve
                
                # Also get relationship files for drawings
                for f in template_zip.namelist():
                    if 'xl/drawings' in f and '.rels' in f:
                        files_to_preserve.append(f)
                    if f == 'xl/worksheets/_rels/sheet1.xml.rels':
                        files_to_preserve.append(f)
                
                if not files_to_preserve:
                    return
                
                # Create temp file
                with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                    temp_path = temp_file.name
                
                # Copy output to temp
                shutil.copy2(output_path, temp_path)
                
                # Add preserved files to output
                with ZipFile(temp_path, 'a', ZIP_DEFLATED) as output_zip:
                    for file_path in files_to_preserve:
                        try:
                            if file_path not in output_zip.namelist():
                                file_data = template_zip.read(file_path)
                                output_zip.writestr(file_path, file_data)
                        except Exception as e:
                            print(f"Could not preserve {file_path}: {e}")
                
                # Copy temp back to output
                shutil.copy2(temp_path, output_path)
                
                # Clean up temp
                try:
                    os.remove(temp_path)
                except:
                    pass
                
                print(f"Preserved {len(files_to_preserve)} drawing/image files")
        except Exception as e:
            print(f"Note: Could not preserve images: {e}")
            # Don't fail - continue without images
    
    def generate(self, output_path, company_name, sakadastro, address, invoice_number, changes=None, items=None):
        """
        Generate a new Excel file based on template with specified changes.

        Args:
            output_path (str): Path where the new Excel file will be saved
            company_name (str): Company name for cell A12 (will be prefixed with "კომპ/სახელი")
            sakadastro (str): Sakadastro value for cell A13 (will be prefixed with "ს/კ")
            address (str): Address for cell A14 (will be prefixed with "მისამართი")
            invoice_number (str/int): Invoice number for cell D5
            changes (dict): Additional cell changes (optional)
                          Example: {'A1': 'New Value', 'B2': 123, 'C3': 45.67}
            items (list): Optional items/types for rows 17-24
                         Each item is a dict with keys: 'type', 'quantity', 'price'
                         Example: [{'type': 'Service A', 'quantity': 2, 'price': 100},
                                  {'type': 'Service B', 'quantity': 1, 'price': 50}]

        Returns:
            str: Path to the generated file
        """
        if changes is None:
            changes = {}
        if items is None:
            items = []

        # Prepend prefixes to the required fields
        company_name = f"კომპ/სახელი {company_name}" if company_name else "კომპ/სახელი"
        sakadastro = f"ს/კ {sakadastro}" if sakadastro else "ს/კ"
        address = f"მისამართი {address}" if address else "მისამართი"

        # Create output directory if it doesn't exist
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)

        # IMPORTANT: Copy the entire template file first to preserve images, formatting, and all content
        shutil.copy2(self.template_path, output_path)

        # Modify the copied file directly by extracting, modifying XML, and re-zipping
        # This preserves ALL content including images and their references
        self._modify_excel_in_place(output_path, company_name, sakadastro, address, invoice_number, changes, items)

        print(f"Excel file generated: {output_path}")

        # Clean up temporary file if it was created
        if hasattr(self, '_temp_file') and os.path.exists(self._temp_file):
            try:
                os.remove(self._temp_file)
            except:
                pass

        return output_path
    
    def generate_pdf(self, excel_path, pdf_path=None):
        """
        Convert an Excel file to PDF using LibreOffice.
        
        Args:
            excel_path (str): Path to the Excel file to convert
            pdf_path (str): Optional path for the PDF file. If not provided, 
                           will use same name as Excel file with .pdf extension
        
        Returns:
            str: Path to the generated PDF file
        """
        if not os.path.exists(excel_path):
            raise FileNotFoundError(f"Excel file not found: {excel_path}")
        
        if pdf_path is None:
            # Generate PDF path from Excel path
            pdf_path = os.path.splitext(excel_path)[0] + '.pdf'
        
        try:
            # Use LibreOffice to convert Excel to PDF with explicit image inclusion
            # --headless: Run without GUI
            # --convert-to pdf: Convert to PDF with default settings
            # --outdir: Output directory
            
            output_dir = os.path.dirname(pdf_path) or '.'
            excel_abs_path = os.path.abspath(excel_path)
            
            # Ensure output directory exists
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            command = [
                '/snap/bin/libreoffice',
                '--headless',
                '--norestore',  # Don't restore previous session
                '--convert-to', 'pdf',
                '--outdir', output_dir,
                excel_abs_path
            ]
            
            # Try standard libreoffice if snap version not available
            try:
                result = subprocess.run(command, check=True, capture_output=True, timeout=60)
            except FileNotFoundError:
                command[0] = 'libreoffice'
                result = subprocess.run(command, check=True, capture_output=True, timeout=60)
            except subprocess.TimeoutExpired:
                raise Exception("LibreOffice PDF conversion timed out after 60 seconds")
            
            # Verify PDF was created
            if not os.path.exists(pdf_path):
                raise Exception(f"PDF file was not created at {pdf_path}")
            
            print(f"PDF file generated: {pdf_path}")
            return pdf_path
        
        except Exception as e:
            print(f"Error generating PDF: {e}")
            print("Make sure LibreOffice is installed: sudo apt install libreoffice")
            raise
    
    def generate_multiple(self, output_dir, changes_list):
        """
        Generate multiple Excel files from the template.
        
        Args:
            output_dir (str): Directory where files will be saved
            changes_list (list): List of tuples (filename, company_name, sakadastro, address, invoice_number, items, additional_changes_dict)
                                Example: [('file1.xlsx', 'Company A', 'Sak001', 'Address 1', 'INV001', 
                                          [{'type': 'Item1', 'quantity': 2, 'price': 100}], {}), 
                                         ('file2.xlsx', 'Company B', 'Sak002', 'Address 2', 'INV002', [], {})]
        
        Returns:
            list: Paths to all generated files
        """
        generated_files = []
        
        for item in changes_list:
            if len(item) == 5:
                filename, company_name, sakadastro, address, invoice_number = item
                items = []
                additional_changes = {}
            elif len(item) == 6:
                filename, company_name, sakadastro, address, invoice_number, items_or_changes = item
                if isinstance(items_or_changes, list):
                    items = items_or_changes
                    additional_changes = {}
                else:
                    items = []
                    additional_changes = items_or_changes
            else:
                filename, company_name, sakadastro, address, invoice_number, items, additional_changes = item
            
            output_path = os.path.join(output_dir, filename)
            self.generate(output_path, company_name, sakadastro, address, invoice_number, additional_changes, items)
            generated_files.append(output_path)
        
        return generated_files


# Example usage
if __name__ == "__main__":
    template_path = "template.xlsx"
    
    # Example 1: Generate single file with required fields only
    generator = ExcelTemplateGenerator(template_path)
    
    generator.generate(
        "output1.xlsx",
        company_name="ACME Corp",
        sakadastro="SAK-2024-001",
        address="123 Main Street, City",
        invoice_number="INV-001"
    )
    
    # Example 2: Generate with items (type, quantity, price)
    items = [
        {'type': 'Service A', 'quantity': 2, 'price': 100},
        {'type': 'Service B', 'quantity': 1, 'price': 50},
    ]
    
    generator.generate(
        "output2.xlsx",
        company_name="Tech Solutions",
        sakadastro="SAK-2024-002",
        address="456 Tech Avenue",
        invoice_number="INV-002",
        items=items
    )
    
    # Example 3: Generate multiple files with items
    changes_list = [
        ('output3.xlsx', 'Company A', 'SAK-001', 'Address A', 'INV-003',
         [{'type': 'Item1', 'quantity': 5, 'price': 75}]),
        ('output4.xlsx', 'Company B', 'SAK-002', 'Address B', 'INV-004',
         [{'type': 'Item2', 'quantity': 3, 'price': 50},
          {'type': 'Item3', 'quantity': 2, 'price': 100}]),
    ]
    
    generator.generate_multiple(".", changes_list)